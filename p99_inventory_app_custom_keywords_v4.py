import re
import threading
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Calibri', color='FFFFFF', bold=True)
NORMAL_FONT = Font(name='Calibri', color='000000')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
HYPER = Font(name='Calibri', color='0563C1', underline='single')

TABLE_STYLE = TableStyleInfo(
    name='TableStyleMedium2',
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
HIGHLIGHT_FONT = Font(name='Calibri', color='9C5700', bold=True)

DEFAULT_KEYWORDS = ['Crown', 'Throne', 'Squire', 'Knight']

EXACT_LINKS = {
    "Helssen's Prismatic Trinket": "https://wiki.project1999.com/Helssen's_Prismatic_Trinket",
    "Crystal Crown of Confusion": "https://wiki.project1999.com/Crystal_Crown_of_Confusion",
}

def slug(name):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(name)).strip("_")

def wiki_link(name):
    if str(name).strip().lower() == "empty":
        return ""
    return EXACT_LINKS.get(name, f"https://wiki.project1999.com/{slug(name)}")

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31] or "Sheet"

def parse_inventory_file(path):
    rows = []
    text = Path(path).read_text(encoding="utf-8", errors="ignore")
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) < 5 or parts[0] == "Location":
            continue
        loc, name, item_id, count, slots = parts[:5]
        try:
            rows.append([loc, name, int(item_id), int(count), int(slots), wiki_link(name)])
        except:
            continue
    return rows

def unique_table_name(wb, base="Table"):
    existing = set()
    for ws in wb.worksheets:
        existing.update(ws.tables.keys())
    i = 1
    while f"{base}_{i}" in existing:
        i += 1
    return f"{base}_{i}"

def row_matches_keywords(name, keywords):
    n = str(name).lower()
    return any(str(k).strip().lower() in n for k in keywords if str(k).strip())

def remove_sheet_if_exists(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

def add_name_highlights(ws, keywords, name_col="B"):
    if ws.max_row < 2:
        return

    dxf = DifferentialStyle(fill=HIGHLIGHT_FILL, font=HIGHLIGHT_FONT)
    end_col = get_column_letter(ws.max_column)
    target_range = f"A2:{end_col}{ws.max_row}"

    for word in keywords:
        w = str(word).strip()
        if not w:
            continue
        safe = w.replace('"', '""')
        rule = Rule(type="expression", dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{safe}",${name_col}2)))']
        ws.conditional_formatting.add(target_range, rule)

def format_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

def format_inventory_sheet(ws, wb, keywords):
    format_header(ws)

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font 