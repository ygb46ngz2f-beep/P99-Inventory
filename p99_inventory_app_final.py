import re
import threading
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Calibri', color='FFFFFF', bold=True)
NORMAL_FONT = Font(name='Calibri', color='000000')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
HYPER = Font(name='Calibri', color='0563C1', underline='single')
TABLE_STYLE = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFF2CC')
HIGHLIGHT_FONT = Font(name='Calibri', color='9C5700', bold=True)
DEFAULT_KEYWORDS = ['Crown', 'Throne', 'Squire', 'Knight']
EXACT_LINKS = {
    "Helssen's Prismatic Trinket": "https://wiki.project1999.com/Helssen's_Prismatic_Trinket",
    "Crystal Crown of Confusion": "https://wiki.project1999.com/Crystal_Crown_of_Confusion",
}

def slug(name):
    return re.sub(r"[^A-Za-z0-9]+", '_', str(name)).strip('_')

def wiki_link(name):
    if str(name).strip().lower() == 'empty':
        return ''
    return EXACT_LINKS.get(name, f'https://wiki.project1999.com/{slug(name)}')

def clean_sheet_name(name):
    return re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31] or 'Sheet'

def parse_inventory_file(path):
    rows = []
    text = Path(path).read_text(encoding='utf-8', errors='ignore')
    for line in text.splitlines():
        parts = line.split('\t')
        if len(parts) < 5 or parts[0] == 'Location':
            continue
        loc, name, item_id, count, slots = parts[:5]
        try:
            rows.append([loc, name, int(item_id), int(count), int(slots), wiki_link(name)])
        except:
            continue
    return rows

def unique_table_name(wb, base='Table'):
    existing = set()
    for ws in wb.worksheets:
        existing.update(ws.tables.keys())
    i = 1
    while f'{base}_{i}' in existing:
        i += 1
    return f'{base}_{i}'

def row_matches_keywords(name, keywords):
    n = str(name).lower()
    return any(str(k).strip().lower() in n for k in keywords if str(k).strip())

def remove_sheet_if_exists(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

def add_name_highlights(ws, keywords, name_col='B'):
    if ws.max_row < 2:
        return
    dxf = DifferentialStyle(fill=HIGHLIGHT_FILL, font=HIGHLIGHT_FONT)
    end_col = get_column_letter(ws.max_column)
    target_range = f'A2:{end_col}{ws.max_row}'
    for word in keywords:
        w = str(word).strip()
        if not w:
            continue
        safe = w.replace('"', '""')
        rule = Rule(type='expression', dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("{safe}",${name_col}2)))']
        ws.conditional_formatting.add(target_range, rule)

def format_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

def format_inventory_sheet(ws, wb, keywords):
    format_header(ws)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = NORMAL_FONT if col != 6 else HYPER
            cell.alignment = CENTER if col in (3, 4, 5) else LEFT
        if ws.cell(row, 6).value:
            ws.cell(row, 6).hyperlink = ws.cell(row, 6).value
    widths = {1:24, 2:42, 3:12, 4:12, 5:12, 6:55}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    table = Table(displayName=unique_table_name(wb, 'Table'), ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)
    add_name_highlights(ws, keywords, 'B')

def rebuild_summary(wb, keywords):
    remove_sheet_if_exists(wb, 'Item Summary')
    ws = wb.create_sheet('Item Summary')
    grouped = defaultdict(lambda: {'total': 0, 'files': set()})
    for sheet in wb.worksheets:
        if sheet.title == 'Item Summary':
            continue
        headers = [sheet.cell(1, c).value for c in range(1, min(sheet.max_column, 6) + 1)]
        if headers[:5] != ['Location', 'Name', 'ID', 'Count', 'Slots']:
            continue
        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(r, 2).value
            if not name or str(name).strip().lower() == 'empty':
                continue
            item_id = sheet.cell(r, 3).value
            count = sheet.cell(r, 4).value or 0
            link = sheet.cell(r, 6).value or wiki_link(name)
            key = (str(name), item_id, link)
            grouped[key]['total'] += int(count)
            grouped[key]['files'].add(sheet.title)
    rows = []
    for (name, item_id, link), data in grouped.items():
        highlighted = 1 if row_matches_keywords(name, keywords) else 0
        rows.append([name, item_id, data['total'], ', '.join(sorted(data['files'])), link, highlighted])
    rows.sort(key=lambda x: (-x[5], -x[2], x[0].lower()))
    ws.append(['Name', 'ID', 'Total Count', 'Files Found In', 'Wiki Link'])
    for row in rows:
        ws.append(row[:5])
    format_header(ws)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 5).value:
            ws.cell(r, 5).hyperlink = ws.cell(r, 5).value
        ws.cell(r, 5).font = HYPER
        ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 3).alignment = CENTER
    widths = {1:42, 2:12, 3:12, 4:30, 5:55}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    table = Table(displayName=unique_table_name(wb, 'ItemSummaryTable'), ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)
    add_name_highlights(ws, keywords, 'A')

def import_files(input_files, output_file, log, keywords):
    output_path = Path(output_file)
    if output_path.exists():
        wb = load_workbook(output_path)
        log(f'Opened existing workbook: {output_path.name}')
    else:
        wb = Workbook()
        wb.remove(wb.active)
        log(f'Created new workbook: {output_path.name}')
    for file_path in input_files:
        rows = parse_inventory_file(file_path)
        sheet_name = clean_sheet_name(Path(file_path).stem)
        remove_sheet_if_exists(wb, sheet_name)
        ws = wb.create_sheet(sheet_name)
        ws.append(['Location', 'Name', 'ID', 'Count', 'Slots', 'Wiki Link'])
        for row in rows:
            ws.append(row)
        format_inventory_sheet(ws, wb, keywords)
        log(f'Imported {Path(file_path).name} -> sheet {sheet_name}')
    rebuild_summary(wb, keywords)
    wb.save(output_path)
    log(f'Saved workbook: {output_path}')
    return output_path

class App:
    def __init__(self, root):
        self.root = root
        self.root.title('P99 Inventory to Excel')
        self.root.geometry('800x600')
        self.files = []
        self.output_var = tk.StringVar(value=str(Path.home() / 'Desktop' / 'p99_inventory.xlsx'))
        self.keyword_var = tk.StringVar()
        frm = ttk.Frame(root, padding=16)
        frm.pack(fill='both', expand=True)
        ttk.Label(frm, text='Project 1999 Inventory Importer', font=('Calibri', 16, 'bold')).pack(anchor='w')
        ttk.Label(frm, text='Select one or more inventory .txt files, then create or update an Excel workbook.').pack(anchor='w', pady=(4,12))
        btns = ttk.Frame(frm)
        btns.pack(fill='x', pady=(0,10))
        ttk.Button(btns, text='Add TXT Files', command=self.add_files).pack(side='left')
        ttk.Button(btns, text='Clear List', command=self.clear_files).pack(side='left', padx=8)
        self.listbox = tk.Listbox(frm, height=8)
        self.listbox.pack(fill='both', expand=False, pady=(0,12))
        kwfrm = ttk.LabelFrame(frm, text='Highlight keywords', padding=10)
        kwfrm.pack(fill='x', pady=(0,10))
        row1 = ttk.Frame(kwfrm)
        row1.pack(fill='x')
        ttk.Entry(row1, textvariable=self.keyword_var).pack(side='left', fill='x', expand=True)
        ttk.Button(row1, text='Add Keyword', command=self.add_keyword).pack(side='left', padx=(8,0))
        ttk.Button(row1, text='Add Default Keywords', command=self.add_defaults).pack(side='left', padx=(8,0))
        ttk.Button(row1, text='Remove Selected', command=self.remove_keyword).pack(side='left', padx=(8,0))
        self.keyword_list = tk.Listbox(kwfrm, height=6, selectmode='extended')
        self.keyword_list.pack(fill='x', pady=(8,0))
        self.add_defaults()
        outfrm = ttk.Frame(frm)
        outfrm.pack(fill='x', pady=(0,10))
        ttk.Label(outfrm, text='Output Excel file:').pack(anchor='w')
        entfrm = ttk.Frame(outfrm)
        entfrm.pack(fill='x', pady=(4,0))
        ttk.Entry(entfrm, textvariable=self.output_var).pack(side='left', fill='x', expand=True)
        ttk.Button(entfrm, text='Browse', command=self.pick_output).pack(side='left', padx=(8,0))
        ttk.Button(frm, text='Create / Update Workbook', command=self.run).pack(anchor='w', pady=(6,10))
        ttk.Label(frm, text='Log:').pack(anchor='w')
        self.logbox = tk.Text(frm, height=14, wrap='word')
        self.logbox.pack(fill='both', expand=True)
    def log(self, msg):
        self.logbox.insert('end', msg + '\n')
        self.logbox.see('end')
        self.root.update_idletasks()
    def add_keyword(self):
        val = self.keyword_var.get().strip()
        if not val:
            return
        existing = list(self.keyword_list.get(0, 'end'))
        if val.lower() not in [x.lower() for x in existing]:
            self.keyword_list.insert('end', val)
        self.keyword_var.set('')
    def add_defaults(self):
        existing = [x.lower() for x in self.keyword_list.get(0, 'end')]
        for kw in DEFAULT_KEYWORDS:
            if kw.lower() not in existing:
                self.keyword_list.insert('end', kw)
    def remove_keyword(self):
        selected = list(self.keyword_list.curselection())
        selected.reverse()
        for i in selected:
            self.keyword_list.delete(i)
    def get_keywords(self):
        return [x for x in self.keyword_list.get(0, 'end') if str(x).strip()]
    def add_files(self):
        paths = filedialog.askopenfilenames(filetypes=[('Text files', '*.txt'), ('All files', '*.*')])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.listbox.insert('end', p)
    def clear_files(self):
        self.files = []
        self.listbox.delete(0, 'end')
    def pick_output(self):
        path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Workbook', '*.xlsx')], initialfile=Path(self.output_var.get()).name)
        if path:
            self.output_var.set(path)
    def run(self):
        if not self.files:
            messagebox.showwarning('No files', 'Please add at least one inventory txt file.')
            return
        output = self.output_var.get().strip()
        if not output:
            messagebox.showwarning('No output', 'Please choose an output Excel file.')
            return
        def job():
            try:
                out = import_files(self.files, output, self.log, self.get_keywords())
                messagebox.showinfo('Done', f'Workbook created/updated:\n{out}')
            except Exception as e:
                self.log(traceback.format_exc())
                messagebox.showerror('Error', str(e))
        threading.Thread(target=job, daemon=True).start()

if __name__ == '__main__':
    root = tk.Tk()
    App(root)
    root.mainloop()
